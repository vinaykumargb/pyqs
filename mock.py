import os
import asyncio
import time
from datetime import datetime
import pytz
from openpyxl import load_workbook
from telegram import Bot, Poll
from telegram.error import TelegramError, TimedOut, NetworkError
import random
from main import post_mcqs_to_telegram

# --------------------------------------
# CONFIGURATION
# --------------------------------------
TELEGRAM_BOT_TOKEN = os.environ.get("TOKEN")
CHAT_ID = -1003018799293
MESSAGE_THREAD_ID = 3
EXCEL_FILE_PATH = "pyqs.xlsx"

# Retry configuration
MAX_RETRIES = 3
RETRY_DELAY = 5  # seconds
REQUEST_TIMEOUT = 30  # seconds

# Updated schedule: Monday (0) to Saturday (5) only
SCHEDULE = {
    0: ["Polity"],
    1: ["Economy"],
    2: ["Ancient India", "Art & Culture", "Medieval India", "Modern India"],
    3: ["Geography"],
    4: ["ScienceAndTech"],
    5: ["Environment"],
    # Sunday (6) is removed
}

IST = pytz.timezone('Asia/Kolkata')

# --------------------------------------
# RETRY DECORATOR
# --------------------------------------
async def retry_on_timeout(func, *args, max_retries=MAX_RETRIES, delay=RETRY_DELAY, **kwargs):
    """Retry a function if it times out"""
    for attempt in range(max_retries):
        try:
            return await func(*args, **kwargs)
        except (TimedOut, NetworkError, asyncio.TimeoutError) as e:
            if attempt < max_retries - 1:
                wait_time = delay * (attempt + 1)  # Exponential backoff
                print(f"Timeout on attempt {attempt + 1}/{max_retries}. Retrying in {wait_time}s... Error: {e}")
                await asyncio.sleep(wait_time)
            else:
                print(f"Failed after {max_retries} attempts: {e}")
                raise
        except Exception as e:
            print(f"Non-timeout error: {e}")
            raise

# --------------------------------------
# CLEAN NEWLINES
# --------------------------------------
def normalize_newlines(text: str) -> str:
    if not text:
        return ""
    return text.replace("\\n", "\n\n")

# --------------------------------------
# EXCEL LOADER
# --------------------------------------
def load_mcqs_from_excel(file_path):
    try:
        workbook = load_workbook(file_path, read_only=True)
        sheet = workbook.active
        headers = [cell.value.strip() if cell.value else "" for cell in sheet[1]]

        mcqs = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            mcq = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            mcqs.append(mcq)

        workbook.close()
        print(f"Loaded {len(mcqs)} MCQs")
        return mcqs

    except Exception as e:
        print(f"Error loading Excel: {e}")
        return []

# --------------------------------------
# SUBJECT-WISE SELECTION
# --------------------------------------
def select_mcqs_for_day(mcqs, day_of_week):
    subjects = SCHEDULE.get(day_of_week, [])
    if not subjects or not mcqs:
        return []

    selected = []

    if day_of_week == 2:  # History special logic
        ancient_art = [m for m in mcqs if m.get('Subject') in ['Ancient India', 'Art & Culture']]
        medieval = [m for m in mcqs if m.get('Subject') == 'Medieval India']
        modern = [m for m in mcqs if m.get('Subject') == 'Modern India']

        if ancient_art:
            selected.append(random.choice(ancient_art))
        if medieval:
            selected.append(random.choice(medieval))
        if len(modern) >= 3:
            selected.extend(random.sample(modern, 3))
        else:
            selected.extend(modern)

        if len(selected) < 5:
            rest = [m for m in mcqs if m.get("Subject") in subjects and m not in selected]
            needed = 5 - len(selected)
            selected.extend(random.sample(rest, min(needed, len(rest))))

    else:
        pool = [m for m in mcqs if m.get("Subject") in subjects]
        selected = random.sample(pool, min(5, len(pool)))

    return selected

# --------------------------------------
# SEND TELEGRAM POLL WITH RETRY
# --------------------------------------
async def post_mcq_poll(bot, mcq, mcq_number):
    try:
        # YEAR CLEANUP
        year_raw = mcq.get('Year')
        try:
            year = int(float(year_raw)) if year_raw else ""
        except:
            year = ""

        year_text = f"\n\n[UPSC CSE Prelims [{year}]]" if year else ""

        # QUESTION
        question_raw = normalize_newlines(str(mcq.get('Question', '')))

        # SUBJECT TAG ONLY FOR Q1
        day_subject = ""
        if mcq_number == 1:
            subject = str(mcq.get("Subject", "")).strip()
            history_subjects = ["Ancient India", "Art & Culture", "Medieval India", "Modern India"]

            if subject in history_subjects:
                day_subject = "#History\n\n"
            elif subject:
                day_subject = f"#{subject}\n\n"

        question_text = f"{day_subject}📝 MCQ {mcq_number}/5\n\n{question_raw}{year_text}"

        # Send message with retry
        await retry_on_timeout(
            bot.send_message,
            chat_id=CHAT_ID,
            text=question_text,
            message_thread_id=MESSAGE_THREAD_ID,
            parse_mode="Markdown",
            read_timeout=REQUEST_TIMEOUT,
            write_timeout=REQUEST_TIMEOUT,
            connect_timeout=REQUEST_TIMEOUT
        )

        # OPTIONS
        options = [
            normalize_newlines(str(mcq.get("Option A", ""))),
            normalize_newlines(str(mcq.get("Option B", ""))),
            normalize_newlines(str(mcq.get("Option C", ""))),
            normalize_newlines(str(mcq.get("Option D", "")))
        ]

        correct_answer = str(mcq.get("Correct Answer", "")).strip().upper()
        correct_index = {"A": 0, "B": 1, "C": 2, "D": 3}.get(correct_answer, 0)

        # EXPLANATION
        explanation_text = normalize_newlines(str(mcq.get("Explanation", "")))

        # Send poll with retry
        await retry_on_timeout(
            bot.send_poll,
            chat_id=CHAT_ID,
            question="Select your answer:",
            options=options,
            type=Poll.QUIZ,
            correct_option_id=correct_index,
            explanation=explanation_text,
            is_anonymous=True,
            message_thread_id=MESSAGE_THREAD_ID,
            read_timeout=REQUEST_TIMEOUT,
            write_timeout=REQUEST_TIMEOUT,
            connect_timeout=REQUEST_TIMEOUT
        )

        print(f"Posted MCQ {mcq_number}")
        await asyncio.sleep(2)

    except Exception as e:
        print(f"Error posting MCQ {mcq_number}: {e}")
        raise  # Re-raise to handle in the caller

# --------------------------------------
# POST DAILY 5 MCQs
# --------------------------------------
async def post_daily_mcqs():
    # Check if today is Sunday (6 = Sunday)
    day = datetime.now(IST).weekday()
    if day == 6:
        print("Today is Sunday. Bot will not run.")
        return
    
    # Initialize bot with custom timeout settings
    bot = Bot(token=TELEGRAM_BOT_TOKEN)
    
    mcqs = load_mcqs_from_excel(EXCEL_FILE_PATH)

    if not mcqs:
        print("No MCQs found")
        return

    selected = select_mcqs_for_day(mcqs, day)

    if not selected:
        print("No MCQs for today's subjects")
        return

    # POST MCQs 1 to 5 with error handling
    successful_posts = 0
    for i, mcq in enumerate(selected, 1):
        try:
            await post_mcq_poll(bot, mcq, i)
            successful_posts += 1
        except Exception as e:
            print(f"Failed to post MCQ {i} after all retries: {e}")
            # Continue with next MCQ instead of stopping

    print(f"Successfully posted {successful_posts}/{len(selected)} MCQs")

    # THEMES SUMMARY
    if successful_posts > 0:
        themes_text = ""
        try:
            themes_text = "📌 Today's PYQ Themes:\n"
            for i, mcq in enumerate(selected, 1):
                topic = mcq.get("Topic", "No topic provided")
                themes_text += f"{i}. {topic}\n"

            await retry_on_timeout(
                bot.send_message,
                chat_id=CHAT_ID,
                text=themes_text,
                message_thread_id=MESSAGE_THREAD_ID,
                read_timeout=REQUEST_TIMEOUT,
                write_timeout=REQUEST_TIMEOUT,
                connect_timeout=REQUEST_TIMEOUT
            )
            
            print("Posted PYQ themes")

        except Exception as e:
            print(f"Error posting themes: {e}")
            return  # Don't proceed to mock MCQs if themes fail
        
        # POST MOCK MCQs using Gemini AI
        try:
            print("\nGenerating and posting Mock MCQs...")
            await asyncio.sleep(3)
            
            # Call the imported function to generate and post mock MCQs
            # Note: This function should also implement retry logic internally
            post_mcqs_to_telegram(CHAT_ID, themes_text, MESSAGE_THREAD_ID)
            
            print("Mock MCQs posted successfully!")
            
        except Exception as e:
            print(f"Error posting mock MCQs: {e}")

# --------------------------------------
# MAIN
# --------------------------------------
if __name__ == "__main__":
    if not TELEGRAM_BOT_TOKEN:
        print("TOKEN missing")
        exit()

    print("Starting bot...")
    try:
        asyncio.run(post_daily_mcqs())
        print("Done!")
    except KeyboardInterrupt:
        print("\nBot stopped by user")
    except Exception as e:
        print(f"Fatal error: {e}")
        exit(1)